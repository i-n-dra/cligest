import base64, unicodedata
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from django.db.models.query import QuerySet
from django.db.models.base import Model
from .aes import AES
from pathlib import Path
from openpyxl.worksheet.table import Table
import ctypes.wintypes

subtitle_font = Font(
    name='Century Gothic',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='203764'
)
center_align = Alignment(
    horizontal='center',
    vertical='center',
    indent=0
)
thick_border = Border(
    right=Side(border_style='thin', color='808080'),
)
ficha_border = Border(
    left=Side(
        border_style='thin',
        color='808080'
    ),
    right=Side(
        border_style='thin',
        color='808080'
    ),
    bottom=Side(
        border_style='thin',
        color='808080'
    ),
    top=Side(
        border_style='thin',
        color='808080'
    )
)

def create_file(wb=Workbook):
    # grab the active worksheet
    ws = wb.active
    ws.title = "CLAVES"
    ws.freeze_panes = 'A2' # freeze top rows

    cat1 = ws['E1:E1']
    for row in cat1:
        for cell in row:
            cell.border = thick_border

    subtitles = [
        'RUN Representante Legal',
        'RUT/RUN Empresa',
        'Única',
        'S.I.I.',
        'Factura Electrónica',
        'Dirección de Trabajo',
    ]

    subt_index = 0
    subtitle_cells = ws['A1:F1']
    for row in subtitle_cells:
        for c in row:
            c.value = subtitles[subt_index]
            c.font = subtitle_font
            c.fill = PatternFill("solid", fgColor="DCE6F1")
            subt_index +=1
    
    return

class exportar_claves_all():
    def get_desktop_path():
                CSIDL_DESKTOP = 0  # escritorio
                SHGFP_TYPE_CURRENT = 0
                buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
                ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
                return Path(buf.value)
    desktop = get_desktop_path() / "Claves.xlsx"
    msg = []
    try:
        wb = load_workbook(str(desktop))
    except FileNotFoundError:
        msg.append('No se encontró el archivo "Claves.xlsx" en el escritorio, se ha creado uno nuevo')
        wb = Workbook(write_only=False)
        create_file(wb)

    ws = wb.active
    ws.sheet_view.showGridLines = False
    def add_row(self, run=str, rut=str, claves=list, n_row=int):

        # agregar claves
        n_cols = 1
        cells_values = [
            run,
            rut,
        ]
        for c in claves:
            cells_values.append(c)
        for cell_value in cells_values:
            self.ws.cell(
                n_row,
                column=n_cols,
                value=cell_value
            ).border = ficha_border
            n_cols += 1
        
        # h/w cols
        for column in self.ws.columns:
            max_length = 0
            col = column[1].column_letter # Get the column letter (e.g., 'A', 'B')
            for cell in column:
                if cell.value is not None:
                    curr_length = len(str(cell.value))
                    cell.alignment = Alignment(wrapText=True, vertical='bottom')
                    if curr_length > max_length:
                        max_length = curr_length

            self.ws.column_dimensions[col].width = max_length + 2

        self.ws.row_dimensions[1].height = 29.25

    def export(self, claves=QuerySet, aes=AES):
        n_row = 1

        for obj in claves:
            n_row += 1
            iv = obj.iv
            run = obj.client.run_rep_legal
            rut = obj.client.run_empresa
            enc_claves = [
                base64.b64decode(obj.unica),
                base64.b64decode(obj.sii),
                base64.b64decode(obj.factura_electronica),
                base64.b64decode(obj.dir_trabajo)
            ]
            dec_claves = []

            for c in enc_claves:
                decrypted = aes.decrypt_cfb(c, iv)
                decrypted = decrypted.decode('utf-8')
                dec_claves.append(decrypted)
            self.add_row(self=exportar_claves_all, run=run, rut=rut, claves=dec_claves, n_row=n_row)

        # tabla
        try:
            t = Table(displayName="Claves", ref=f"A1:F{n_row}")
            self.ws.add_table(t)
        except ValueError:
            pass

        # Save the file
        try:
            self.wb.save(str(self.desktop))
            self.msg.append('Se ha creado "Claves.xlsx" exitosamente')
            return self.msg
        except PermissionError:
            self.msg.append('No se puede escribir el archivo mientras está abierto, por favor, cierre el archivo e intente de nuevo.')
            return self.msg

class exportar_clave(): # en progreso
    def get_desktop_path():
                CSIDL_DESKTOP = 0  # escritorio
                SHGFP_TYPE_CURRENT = 0
                buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
                ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
                return Path(buf.value)
    
    def export(self, client=Model, aes=AES):
        cliente_str = f"Claves{client.client.nombre_rep_legal}{client.client.last_name_1_rep_legal}{client.client.last_name_2_rep_legal}"
        cliente_nor = unicodedata.normalize('NFKD', cliente_str)  
        archivo_for = ''.join([c for c in cliente_nor if not unicodedata.combining(c)])
        archivo = archivo_for+".xlsx"

        desktop = self.get_desktop_path() / archivo
        msg = []
        try:
            wb = load_workbook(str(desktop))
        except FileNotFoundError:
            msg.append(f'No se encontró el archivo "{archivo}" en el escritorio, se ha creado uno nuevo')
            wb = Workbook(write_only=False)
            create_file(wb)

        ws = wb.active
        ws.sheet_view.showGridLines = False
        
        n_row = 2

        iv = client.iv
        run = client.client.run_rep_legal
        rut = client.client.run_empresa
        enc_claves = [
            base64.b64decode(client.unica),
            base64.b64decode(client.sii),
            base64.b64decode(client.factura_electronica),
            base64.b64decode(client.dir_trabajo)
        ]
        dec_claves = []

        for c in enc_claves:
            decrypted = aes.decrypt_cfb(c, iv)
            decrypted = decrypted.decode('utf-8')
            dec_claves.append(decrypted)

        # agregar claves
        n_cols = 1
        cells_values = [
            run,
            rut,
        ]
        for c in dec_claves:
            cells_values.append(c)

        for cell_value in cells_values:
            ws.cell(
                n_row,
                column=n_cols,
                value=cell_value
            ).border = ficha_border
            n_cols += 1
        
        # h/w cols
        for column in ws.columns:
            max_length = 0
            col = column[1].column_letter # Get the column letter (e.g., 'A', 'B')
            for cell in column:
                if cell.value is not None:
                    curr_length = len(str(cell.value))
                    cell.alignment = Alignment(wrapText=True, vertical='bottom')
                    if curr_length > max_length:
                        max_length = curr_length

            ws.column_dimensions[col].width = max_length + 2

        ws.row_dimensions[1].height = 29.25

        # tabla
        try:
            t = Table(displayName="Claves", ref=f"A1:F{n_row}")
            ws.add_table(t)
        except ValueError:
            pass

        # Save the file
        try:
            wb.save(str(desktop))
            msg.append(f'Se ha creado "{archivo}" exitosamente')
            return msg
        except PermissionError:
            msg.append('No se puede escribir el archivo mientras está abierto, por favor, cierre el archivo e intente de nuevo.')
            return msg
