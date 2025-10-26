from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from tempfile import NamedTemporaryFile
from django.http import Http404
from django.db.models.query import QuerySet
from datetime import datetime, time

### copy paste de exportar clientes, wip

titulo_font = Font(
    name='Century Gothic',
    size=18,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='808080'
)
subtitle_font = Font(
    name='Century Gothic',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='203764'
)
center_align = Alignment(
    horizontal='center',
    vertical='center',
    indent=0
)
thick_border = Border(
    right=Side(border_style='thin', color='808080'),
)
ficha_border = Border(
    left=Side(
        border_style='thin',
        color='808080'
    ),
    right=Side(
        border_style='thin',
        color='808080'
    ),
    bottom=Side(
        border_style='thin',
        color='808080'
    ),
    top=Side(
        border_style='thin',
        color='808080'
    )
)

def exportar_pagos_main(clientes=QuerySet, pagos=QuerySet): # 2do arg: pagos=queryset
    path = 'D:/py_venvs/proyectos/proyecto_integracion/cligest/FichasClientes.xlsx'
    msg = []
    try:
        wb = load_workbook(path)
    except FileNotFoundError:
        msg.append('No se encontró el archivo "FichasClientes.xlsx" en el directorio, se ha creado uno nuevo')
        wb = Workbook(write_only=False)
        create_file(wb)

    # grab the active worksheet
    ws = wb.active
    ws.sheet_view.showGridLines = False

    # agregar clientes
    # falta probar con mas de uno xd
    n_row = 3
    while n_row<(clientes.__len__())+3:
        for c in clientes.all():
            print(f'ingresando cliente: {c.id}')

            #a_pagar = pagos.filter(self=pagos,client=c.id)

            reg_tri_list = list(c.reg_tributario.all())
            reg_tri_str = ''
            for r in reg_tri_list:
                reg_tri_str += f'{r}\n'
            reg_tri_str.rstrip('\n')

            gir_ru_list = list(c.giro_rubro.all())
            gir_ru_str = ''
            for r in gir_ru_list:
                gir_ru_str += f'{r}\n'
            gir_ru_str.rstrip('\n')

            cod_sii_list = list(c.codigo_sii.all())
            cod_sii_str = ''
            for r in cod_sii_list:
                cod_sii_str += f'{r}\n'
            cod_sii_str.rstrip('\n')

            tipo_cont_list = list(c.tipo_contabilidad.all())
            tipo_cont_str = ''
            for r in tipo_cont_list:
                tipo_cont_str += f'{r}\n'
            tipo_cont_str.rstrip('\n')

            # rellenando celdas
            ws.cell(
                n_row,
                column = 1,
                value = c.nombre_rep_legal
            ).border = ficha_border

            ws.cell(
                n_row,
                2,
                c.last_name_1_rep_legal
            ).border = ficha_border
            ws.cell(
                n_row,
                3,
                c.last_name_2_rep_legal
            ).border = ficha_border
            ws.cell(
                n_row,
                4,
                c.run_rep_legal
            ).border = ficha_border
            ws.cell(
                n_row,
                5,
                c.tipo_empresa
            ).border = ficha_border
            ws.cell(
                n_row,
                6,
                c.razon_social
            ).border = ficha_border
            ws.cell(
                n_row,
                7,
                c.nombre_fantasia
            ).border = ficha_border
            ws.cell(
                n_row,
                8,
                c.run_empresa
            ).border = ficha_border
            ws.cell(
                n_row,
                9,
                reg_tri_str
            ).border = ficha_border
            ws.cell(
                n_row,
                10,
                gir_ru_str
            ).border = ficha_border
            ws.cell(
                n_row,
                11,
                cod_sii_str
            ).border = ficha_border
            ws.cell(
                n_row,
                12,
                c.n_trabajadores
            ).border = ficha_border
            ws.cell(
                n_row,
                13,
                tipo_cont_str
            ).border = ficha_border
            ws.cell(
                n_row,
                14,
                c.cuenta_corriente
            ).border = ficha_border
            ws.cell(
                n_row,
                15,
                c.n_cuenta_corriente
            ).border = ficha_border
            # ws.cell(
            #     n_row,
            #     16,
            #     a_pagar
            # ).border = ficha_border

            ws.cell(
                n_row,
                17,
                c.email
            ).border = ficha_border
            ws.cell(
                n_row,
                18,
                c.phone_number
            ).border = ficha_border
            ws.cell(
                n_row,
                19,
                str(c.region)
            ).border = ficha_border
            ws.cell(
                n_row,
                20,
                str(c.comuna)
            ).border = ficha_border
            ws.cell(
                n_row,
                21,
                c.address
            ).border = ficha_border

            n_row += 1
            

    for column in ws.columns:
        max_length = 0
        col = column[2].column_letter # Get the column letter (e.g., 'A', 'B')
        for cell in column:
            if cell.value is not None:
                curr_length = len(str(cell.value))
                cell.alignment = Alignment(wrapText=True, vertical='justify')
                if curr_length > max_length:
                    max_length = curr_length

        ws.column_dimensions[col].width = max_length + 2

    ws.row_dimensions[1].height = 24

    # Save the file

    # opción 1
    # with NamedTemporaryFile() as tmp:
    #         wb.save(tmp.name)
    #         tmp.seek(0)
    #         stream = tmp.read()

    # opción 2
    try:
        wb.save('FichasClientes.xlsx') 
        # + guardar una copia en el escritorio?
        msg.append('Se ha creado "FichasClientes.xlsx" exitosamente')
        return msg
    except PermissionError:
        raise Http404('Por favor, cierre el archivo antes de confirmar los cambios')

def create_file(wb=Workbook):
    # grab the active worksheet
    ws = wb.active
    ws.title = "Clientes"
    ws.freeze_panes = 'A3' # freeze top rows

    cat1 = ws['E1:E3']
    for row in cat1:
        for cell in row:
            cell.border = thick_border

    ws['A1'] = 'Datos Representante Legal'
    ws['F1'] = 'Datos Empresa'
    ws['L1'] = 'Datos Financieros'
    ws['Q1'] = 'Datos de Contacto'
    ws['U1'] = 'Otros'
    rep_legal_cell = ws['A1']
    rep_legal_cell.font = titulo_font
    rep_legal_cell.alignment = center_align
    ws.merge_cells('A1:E1')
    emp_cell = ws['F1']
    emp_cell.font = titulo_font
    emp_cell.alignment = center_align
    ws.merge_cells('F1:K1')
    fin_cell = ws['L1']
    fin_cell.font = titulo_font
    fin_cell.alignment = center_align
    ws.merge_cells('L1:P1')
    cont_cell = ws['P1']
    cont_cell.font = titulo_font
    cont_cell.alignment = center_align
    ws.merge_cells('Q1:T1')
    otro_cell = ws['U1']
    otro_cell.font = titulo_font
    otro_cell.alignment = center_align

    # Rows can also be appended
    subtitles = [
        # Datos Representante Legal #
        'Nombre(s)', # preguntar sobre esto ??
        'Apellido Paterno',
        'Apellido Materno',
        'RUN',
        'Tipo de Empresa',
        # Datos Empresa #
        'Razón Social',
        'Nombre de Fantasía',
        'RUT',
        'Régimen Tributario',
        'Giro / Rubro',
        'Código S.I.I.',
        'Trabajadores',
        # Datos Financieros #
        'Tipo de Contabilidad',
        'Cuenta Corriente',
        'N° Cuenta Corriente',
        'Deuda',
        # Datos Contacto #
        'Correo Electrónico',
        'Teléfono / Celular',
        'Región',
        'Comuna',
        'Dirección',
        'Añadido el'
    ]

    subt_index = 0
    subtitle_cells = ws['A2:U2']
    for row in subtitle_cells:
        for c in row:
            c.value = subtitles[subt_index]
            c.font = subtitle_font
            c.fill = PatternFill("solid", fgColor="DCE6F1")
            subt_index +=1
    
    return