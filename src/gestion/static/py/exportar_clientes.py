import unicodedata
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from django.db.models.query import QuerySet
from django.utils import timezone
from openpyxl.worksheet.table import Table
from pathlib import Path
import ctypes.wintypes

titulo_font = Font(
    name='Century Gothic',
    size=16,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='203764'
)
subtitle_font = Font(
    name='Century Gothic',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='203764'
)
center_align = Alignment(
    horizontal='center',
    vertical='center',
    indent=0
)
thick_border = Border(
    right=Side(border_style='thin', color='808080'),
)
ficha_border = Border(
    left=Side(
        border_style='thin',
        color='808080'
    ),
    right=Side(
        border_style='thin',
        color='808080'
    ),
    bottom=Side(
        border_style='thin',
        color='808080'
    ),
    top=Side(
        border_style='thin',
        color='808080'
    )
)

def exportar_clientes_main(clientes=QuerySet, pagos=QuerySet):
    n_row = 3
    n_cli = (clientes.__len__())+3

    def get_desktop_path():
                CSIDL_DESKTOP = 0  # escritorio
                SHGFP_TYPE_CURRENT = 0
                buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
                ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
                return Path(buf.value)
    desktop = get_desktop_path() / "FichasClientes.xlsx"    
    msg = []
    try:
        wb = load_workbook(str(desktop))
    except FileNotFoundError:
        msg.append('No se encontró el archivo "FichasClientes.xlsx" en el escritorio, se ha creado uno nuevo')
        wb = Workbook(write_only=False)
        create_file(wb)

    # grab the active worksheet
    ws = wb.active

    # tabla
    try:
        t = Table(displayName="Clientes", ref=f"A2:U{str(n_cli-1)}")
        ws.add_table(t)
    except ValueError:
        ws.delete_rows(3,n_cli-3)

    ws.sheet_view.showGridLines = False

    # agregar clientes
    while n_row < n_cli:
        for c in clientes.all():
            now = timezone.now()
            a_pagar_exists = pagos.filter(
                client=c.id,
                created_at__year=now.year,
                created_at__month=now.month
            ).exists()
            if a_pagar_exists:
                a_pagar_obj = pagos.filter(
                    client=c.id,
                    created_at__year=now.year,
                    created_at__month=now.month
                ).last()
                a_pagar = a_pagar_obj.a_pagar
            else: a_pagar = 0

            reg_tri_list = list(c.reg_tributario.all())
            reg_tri_str = ''
            for r in reg_tri_list:
                reg_tri_str += f'{r}\n'
            reg_tri_str.rstrip('\n')

            gir_ru_list = list(c.giro_rubro.all())
            gir_ru_str = ''
            for r in gir_ru_list:
                gir_ru_str += f'{r}\n'
            gir_ru_str.rstrip('\n')

            cod_sii_list = list(c.codigo_sii.all())
            cod_sii_str = ''
            for r in cod_sii_list:
                cod_sii_str += f'{r}\n'
            cod_sii_str.rstrip('\n')

            tipo_cont_list = list(c.tipo_contabilidad.all())
            tipo_cont_str = ''
            for r in tipo_cont_list:
                tipo_cont_str += f'{r}\n'
            tipo_cont_str.rstrip('\n')

            # rellenando celdas
            n_cols = 1
            cells_values = [
                c.nombre_rep_legal,
                c.last_name_1_rep_legal,
                c.last_name_2_rep_legal,
                c.run_rep_legal,
                c.tipo_empresa,
                c.razon_social,
                c.nombre_fantasia,
                c.run_empresa,
                reg_tri_str,
                gir_ru_str,
                cod_sii_str,
                c.n_trabajadores,
                tipo_cont_str,
                c.cuenta_corriente,
                c.n_cuenta_corriente,
                a_pagar,
                c.email,
                c.phone_number,
                str(c.region),
                str(c.comuna),
                c.address,
            ]

            for cell_value in cells_values:
                ws.cell(
                    n_row,
                    column = n_cols,
                    value = cell_value
                ).border = ficha_border

                n_cols += 1

            n_row += 1
    
    # h/w cols
    for column in ws.columns:
        max_length = 0
        col = column[2].column_letter # Get the column letter (e.g., 'A', 'B')
        for cell in column:
            if cell.value is not None:
                curr_length = len(str(cell.value))
                cell.alignment = Alignment(wrapText=True, vertical='justify')
                if curr_length > max_length:
                    max_length = curr_length

        ws.column_dimensions[col].width = max_length + 2

    ws.column_dimensions["P"].width = 14
    ws.row_dimensions[1].height = 24

    # currency
    for cell in ws["P"]:
        cell.number_format = "$#,##0"

    # Save the file
    try:
        wb.save(str(desktop))
        msg.append('Se ha creado "FichasClientes.xlsx" exitosamente')
        return msg
    except PermissionError:
        msg.append('No se puede escribir el archivo mientras está abierto, por favor, cierre el archivo e intente de nuevo.')
        return msg

def exportar_cliente(cliente=QuerySet, pagos=QuerySet):
    def get_desktop_path():
                CSIDL_DESKTOP = 0  # escritorio
                SHGFP_TYPE_CURRENT = 0
                buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
                ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
                return Path(buf.value)
    cliente_str = f"FichaDe{cliente.nombre_rep_legal}{cliente.last_name_1_rep_legal}{cliente.last_name_2_rep_legal}"
    cliente_nor = unicodedata.normalize('NFKD', cliente_str)  
    archivo_for = ''.join([c for c in cliente_nor if not unicodedata.combining(c)])
    archivo = archivo_for+".xlsx"

    desktop = get_desktop_path() / archivo
    msg = []
    try:
        wb = load_workbook(str(desktop))
    except FileNotFoundError:
        msg.append(f'No se encontró el archivo "{archivo}" en el escritorio, se ha creado uno nuevo')
        wb = Workbook(write_only=False)
        create_file(wb)

    # grab the active worksheet
    ws = wb.active
    
    ws.sheet_view.showGridLines = False

    # agregar cliente
    n_row = 3
    now = timezone.now()
    a_pagar_exists = pagos.filter(
        client=cliente.id,
        created_at__year=now.year,
        created_at__month=now.month
    ).exists()
    if a_pagar_exists:
        a_pagar_obj = pagos.filter(
            client=cliente.id,
            created_at__year=now.year,
            created_at__month=now.month
        ).last()
        a_pagar = a_pagar_obj.a_pagar
    else: a_pagar = 0

    reg_tri_list = list(cliente.reg_tributario.all())
    reg_tri_str = ''
    for r in reg_tri_list:
        reg_tri_str += f'{r}\n'
    reg_tri_str.rstrip('\n')

    gir_ru_list = list(cliente.giro_rubro.all())
    gir_ru_str = ''
    for r in gir_ru_list:
        gir_ru_str += f'{r}\n'
    gir_ru_str.rstrip('\n')

    cod_sii_list = list(cliente.codigo_sii.all())
    cod_sii_str = ''
    for r in cod_sii_list:
        cod_sii_str += f'{r}\n'
    cod_sii_str.rstrip('\n')

    tipo_cont_list = list(cliente.tipo_contabilidad.all())
    tipo_cont_str = ''
    for r in tipo_cont_list:
        tipo_cont_str += f'{r}\n'
    tipo_cont_str.rstrip('\n')

    # rellenando celdas
    n_cols = 1
    cells_values = [
        cliente.nombre_rep_legal,
        cliente.last_name_1_rep_legal,
        cliente.last_name_2_rep_legal,
        cliente.run_rep_legal,
        cliente.tipo_empresa,
        cliente.razon_social,
        cliente.nombre_fantasia,
        cliente.run_empresa,
        reg_tri_str,
        gir_ru_str,
        cod_sii_str,
        cliente.n_trabajadores,
        tipo_cont_str,
        cliente.cuenta_corriente,
        cliente.n_cuenta_corriente,
        a_pagar,
        cliente.email,
        cliente.phone_number,
        str(cliente.region),
        str(cliente.comuna),
        cliente.address,
    ]

    for cell_value in cells_values:
        ws.cell(
            n_row,
            column = n_cols,
            value = cell_value
        ).border = ficha_border

        n_cols += 1
    
    # h/w cols
    for column in ws.columns:
        max_length = 0
        col = column[2].column_letter # Get the column letter (e.g., 'A', 'B')
        for cell in column:
            if cell.value is not None:
                curr_length = len(str(cell.value))
                cell.alignment = Alignment(wrapText=True, vertical='justify')
                if curr_length > max_length:
                    max_length = curr_length

        ws.column_dimensions[col].width = max_length + 2

    ws.column_dimensions["P"].width = 14
    ws.row_dimensions[1].height = 24

    # tabla
    try:
        t = Table(displayName=f"FichaDe{cliente.nombre_rep_legal}{cliente.last_name_1_rep_legal}{cliente.last_name_2_rep_legal}", ref=f"A2:U3")
        ws.add_table(t)
    except ValueError:
        pass # ya existe una tabla

    # currency
    for cell in ws["P"]:
        cell.number_format = "$#,##0"

    # Save the file
    try:
        wb.save(str(desktop))
        msg.append(f'Se ha creado "{archivo}" exitosamente')
        return msg
    except PermissionError:
        msg.append('No se puede escribir el archivo mientras está abierto, por favor, cierre el archivo e intente de nuevo.')
        return msg

def create_file(wb=Workbook):
    # grab the active worksheet
    ws = wb.active
    ws.title = "CLIENTES"
    ws.freeze_panes = 'A3' # freeze top rows

    cat1 = ws['E1:E3']
    for row in cat1:
        for cell in row:
            cell.border = thick_border

    ws['A1'] = 'DATOS REPRESENTANTE LEGAL'
    ws['F1'] = 'DATOS EMPRESA'
    ws['L1'] = 'DATOS FINANCIEROS'
    ws['Q1'] = 'DATOS DE CONTACTO'
    ws['U1'] = 'OTROS'
    rep_legal_cell = ws['A1']
    rep_legal_cell.font = titulo_font
    rep_legal_cell.alignment = center_align
    ws.merge_cells('A1:E1')
    emp_cell = ws['F1']
    emp_cell.font = titulo_font
    emp_cell.alignment = center_align
    ws.merge_cells('F1:K1')
    fin_cell = ws['L1']
    fin_cell.font = titulo_font
    fin_cell.alignment = center_align
    ws.merge_cells('L1:P1')
    cont_cell = ws['Q1']
    cont_cell.font = titulo_font
    cont_cell.alignment = center_align
    ws.merge_cells('Q1:T1')
    otro_cell = ws['U1']
    otro_cell.font = titulo_font
    otro_cell.alignment = center_align

    subtitles = [
        # Datos Representante Legal #
        'Nombre(s)',
        'Apellido Paterno',
        'Apellido Materno',
        'RUN',
        'Tipo de Empresa',
        # Datos Empresa #
        'Razón Social',
        'Nombre de Fantasía',
        'RUN / RUT',
        'Régimen Tributario',
        'Giro / Rubro',
        'Código S.I.I.',
        'Trabajadores',
        # Datos Financieros #
        'Tipo de Contabilidad',
        'Cuenta Corriente',
        'N° Cuenta Corriente',
        'A Pagar',
        # Datos Contacto #
        'Correo Electrónico',
        'Teléfono / Celular',
        'Región',
        'Comuna',
        'Dirección',
    ]

    subt_index = 0
    subtitle_cells = ws['A2:U2']
    for row in subtitle_cells:
        for c in row:
            c.value = subtitles[subt_index]
            c.font = subtitle_font
            c.fill = PatternFill("solid", fgColor="DCE6F1")
            subt_index +=1
    
    return