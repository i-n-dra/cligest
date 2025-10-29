from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from django.db.models.query import QuerySet
from datetime import datetime, time
from pathlib import Path
from openpyxl.worksheet.table import Table
import ctypes.wintypes

### en progreso

titulo_font = Font(
    name='Century Gothic',
    size=18,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='808080'
)
subtitle_font = Font(
    name='Century Gothic',
    size=11,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='203764'
)
center_align = Alignment(
    horizontal='center',
    vertical='center',
    indent=0
)
thick_border = Border(
    right=Side(border_style='thin', color='808080'),
)
ficha_border = Border(
    left=Side(
        border_style='thin',
        color='808080'
    ),
    right=Side(
        border_style='thin',
        color='808080'
    ),
    bottom=Side(
        border_style='thin',
        color='808080'
    ),
    top=Side(
        border_style='thin',
        color='808080'
    )
)

class exportar_pagos_deuda():
    def create_file(wb=Workbook):
        # grab the active worksheet
        ws = wb.active
        ws.title = "PAGOS"
        ws.freeze_panes = 'A2' # freeze top rows

        cat1 = ws['A1:A2']
        for row in cat1:
            for cell in row:
                cell.border = thick_border

        # Rows can also be appended
        subtitles = [
            'RUN Representante Legal',
            'RUT/RUN Empresa',
            'Monto a Pagar',
            'Última modificación'
        ]

        subt_index = 0
        subtitle_cells = ws['A1:D1']
        for row in subtitle_cells:
            for c in row:
                c.value = subtitles[subt_index]
                c.font = subtitle_font
                c.fill = PatternFill("solid", fgColor="DCE6F1")
                subt_index +=1
        
        return

    def get_desktop_path():
                CSIDL_DESKTOP = 0  # escritorio
                SHGFP_TYPE_CURRENT = 0
                buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
                ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
                return Path(buf.value)
    desktop = get_desktop_path() / "Deudas.xlsx"
    msg = []
    try:
        wb = load_workbook(str(desktop))
    except FileNotFoundError:
        msg.append('No se encontró el archivo "Deudas.xlsx" en el directorio, se ha creado uno nuevo')
        wb = Workbook(write_only=False)
        create_file(wb)

    ws = wb.active
    ws.sheet_view.showGridLines = False

    def add_row(self, deudas=QuerySet):

        # agregar deudas
        n_row = 2
        n_deu = (deudas.__len__())+2
        for d in deudas:
            cells_values = [
                    d.nombre_rep_legal,
                    d.last_name_1_rep_legal,
                ]
            n_cols = 1
            for cell_value in cells_values:
                self.ws.cell(
                    n_row,
                    column=n_cols,
                    value=cell_value
                ).border = ficha_border
                n_cols += 1
        
        # h/w cols
        for column in self.ws.columns:
            max_length = 0
            col = column[1].column_letter # Get the column letter (e.g., 'A', 'B')
            for cell in column:
                if cell.value is not None:
                    curr_length = len(str(cell.value))
                    cell.alignment = Alignment(wrapText=True, vertical='bottom')
                    if curr_length > max_length:
                        max_length = curr_length

            self.ws.column_dimensions[col].width = max_length + 2

        self.ws.row_dimensions[1].height = 29.25

        # Save the file
        try:
            self.wb.save(str(self.desktop))
            self.msg.append('Se ha creado "Deudas.xlsx" exitosamente')
            return self.msg
        except PermissionError:
            self.msg.append('No se puede escribir el archivo mientras está abierto, por favor, cierre el archivo e intente de nuevo.')
            return self.msg

        print(f"cliente (run/rut): {run}, {rut}\npagos: {pagos}")

    def export(self, deudas=QuerySet):
        msg = []
        n_row = 2

        for obj in deudas:
            run = obj.client.run_rep_legal
            rut = obj.client.run_empresa


            self.add_row(run, rut, n_row)
            n_row += 1

        # tabla
        try:
            t = Table(displayName="Deudas", ref=f"A1:F{n_row}")
            self.ws.add_table(t)
        except ValueError:
            pass

        # Save the file
        try:
            self.wb.save(str(self.desktop))
            self.msg.append('Se ha creado "Deudas.xlsx" exitosamente')
            return self.msg
        except PermissionError:
            self.msg.append('No se puede escribir el archivo mientras está abierto, por favor, cierre el archivo e intente de nuevo.')
            return self.msg


############################################


class exportar_pagos_historial():
    def create_file(wb=Workbook):
        # grab the active worksheet
        ws = wb.active
        ws.title = "HISTORIAL"
        ws.freeze_panes = 'A2' # freeze top rows

        cat1 = ws['A1:A2']
        for row in cat1:
            for cell in row:
                cell.border = thick_border

        # Rows can also be appended
        subtitles = [
            'RUN Representante Legal',
            'RUT/RUN Empresa',
            'Monto a Pagar',
            'Última modificación'
        ]

        subt_index = 0
        subtitle_cells = ws['A2:D2']
        for row in subtitle_cells:
            for c in row:
                c.value = subtitles[subt_index]
                c.font = subtitle_font
                c.fill = PatternFill("solid", fgColor="DCE6F1")
                subt_index +=1
        
        return
    
    def get_desktop_path():
                CSIDL_DESKTOP = 0  # escritorio
                SHGFP_TYPE_CURRENT = 0
                buf = ctypes.create_unicode_buffer(ctypes.wintypes.MAX_PATH)
                ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
                return Path(buf.value)
    desktop = get_desktop_path() / "DeclaracionesHistorial.xlsx"
    msg = []
    try:
        wb = load_workbook(str(desktop))
    except FileNotFoundError:
        msg.append('No se encontró el archivo "DeclaracionesHistorial.xlsx" en el directorio, se ha creado uno nuevo')
        wb = Workbook(write_only=False)
        create_file(wb)

    ws = wb.active
    ws.sheet_view.showGridLines = False

    def export(self, pagos=QuerySet):
        msg = []
        n_row = 2

        for obj in pagos:
            run = obj.client.run_rep_legal
            rut = obj.client.run_empresa


            self.add_row(run, rut, n_row)
            n_row += 1

        # tabla
        try:
            t = Table(displayName="Historial", ref=f"A1:F{n_row}")
            self.ws.add_table(t)
        except ValueError:
            pass

        # Save the file
        try:
            self.wb.save(str(self.desktop))
            self.msg.append('Se ha creado "DeclaracionesHistorial.xlsx" exitosamente')
            return self.msg
        except PermissionError:
            self.msg.append('No se puede escribir el archivo mientras está abierto, por favor, cierre el archivo e intente de nuevo.')
            return self.msg